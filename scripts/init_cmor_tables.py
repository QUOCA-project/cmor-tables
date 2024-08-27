#!/usr/bin/env python
'''
Attempt to match QUOCA requested variables as listed in protocol with variables in CMOR tables of existing projects.

QUOCA requested variables are obtained from spreadsheet provided by Alison.
Need to download it and copy to the working dir.

Usage example (by James on ECCC hpcr-vis machine):
    load_conda  [alias to make conda available on hpcr-vis]
    conda activate rja_search_esgf_v1  [conda env that happens to have the needed modules]
    ipython   [don't need to use ipython, could just run at command line: python init_cmor_tables.py]
    run -i init_cmor_tables.py

Needs access to various projects' CMOR tables as indicated below in 'projects' dict.
'''

import os
import json
import urllib.request
import openpyxl as xp
import datetime

# Dict 'projects' will be populated with info about CMOR tables for each project
projects = {
    'QBOi' : {
        # https://gitlab.com/JamesAnstey/qboi-cmor-tables
        'path' : 'qboi-cmor-tables/Tables',
        'exclude tables' : [],
        },
   'CCMI2022' : {
        # https://github.com/cedadev/ccmi-2022
        'path' : 'ccmi-2022/Tables',
        'exclude tables' : [],
    },
    'SNAPSI' : {
        # https://github.com/cedadev/snap
        'path' : 'snapsi-cmor-tables/Tables',
        'exclude tables' : ['6hrRef'],
    },
    'CMIP6' : {
        # https://github.com/PCMDI/cmip6-cmor-tables
        'path' : 'cmip6-cmor-tables/Tables',
        'exclude tables' : [],
    },
}

###############################################################################
def set_variable_short_name(variable_id) -> str:
    orig_vid = variable_id
    vid = orig_vid.lower()
    rename = {
        'w*' : 'wtem',
        'v*' : 'vtem',
        'psi*' : 'psitem',
        'wa' : 'wap', # let's assume wa (m s-1) can be wap (Pa s-1)
    }
    if vid in rename:
        # if variable is a specific case
        vid = rename[vid]
    else:
        # if variable is not a specific case, apply general rules
        replace = {
            ' ' : '-',
            '_' : '-',
            '--' : '-',
            '*' : '',
            '’' : 'p', # prime
        }
        for s in replace:
            if s in vid:  # underscores are not allowed in variable name
                # assert replace[s] not in vid
                vid = vid.replace(s, replace[s])
        vid = vid.lower() # variable name should be lowercase
        if vid != orig_vid:
            print(f'renamed variable: {orig_vid} --> {vid}')
    return vid
###############################################################################
# Get info on QUOCA requested variables from Alison's spreadsheet
spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1iCjMc1772CQpc1QWYSnAU7GnPPf_Fo1FHGzURtvrOac'

# filename = 'QUOCA_data_request_v1.xlsx'  # original version by Alison, downloaded 25 Aug 2024
filename = 'QUOCA_data_request_v2.xlsx'  # with changes added by James, downloaded 26 Aug 2024
path = '.'

get_spreadsheet = False
if get_spreadsheet:
    # oops, this only works if the spreadsheet is readable by all
    url = os.path.join(spreadsheet_url, 'export?exportFormat=xlsx')
    urllib.request.urlretrieve(url, os.path.join(path, filename))
    print('downloaded spreadsheet to ' + filepath)

wb = xp.load_workbook(filename=os.path.join(path, filename))
print('loaded ' + filename)

rename_columns = {
    'variable name' : 'variable_id',
    'temporal resolution' : 'frequency',
    'cf standard name' : 'standard_name',
    'longname [units]' : 'long_name [units]',
    'long name [units]' : 'long_name [units]',
}
rename_freqs = {'Monthly' : 'mon', 'Daily' : 'day', '6-hourly' : '6hrPt'}
protocol_vars = {}
assert len(wb.sheetnames) == len(set(wb.sheetnames))
columns0 = None
for sheet in wb.sheetnames:
    for k,row in enumerate(wb[sheet].rows):
        if k == 0:
            columns = [cell.value.lower() for cell in row]
            assert len(columns) == len(set(columns))
            for old_name, new_name in rename_columns.items():
                if old_name in columns:
                    columns[columns.index(old_name)] = new_name
        else:
            var_info = dict(zip(columns, [cell.value for cell in row]))
            # if (var_info['variable name'] is None) and (var_info['cf standard name'] is None):
            if (var_info['variable_id'] is None) or var_info['variable_id'].startswith(sheet):
                continue
            var_info['variable name in spreadsheet'] = var_info['variable_id']
            var_info['variable_id'] = set_variable_short_name(var_info['variable_id'])
            freqs = [s.strip() for s in var_info['frequency'].split(',')]
            for freq in freqs:
                var_info = dict(var_info)
                freq = rename_freqs[freq]
                var_info.update({'frequency' : freq})
                key = (sheet, freq, var_info['variable_id'])
                assert key not in protocol_vars
                protocol_vars[key] = var_info
    n = len([key for key in protocol_vars if key[0] == sheet])
    print(f'Found {n} variables for table {sheet}')
n = len(protocol_vars)
print(f'--> Found total {n} variables in spreadsheet {filename}\n')

# which units to use? 'units' and units in 'long name' arean't always consistent
replace_units_with_long_name_units = True
if replace_units_with_long_name_units:
    for key, var_info in protocol_vars.items():
        assert 'long_name' not in var_info
        for sep in ['(', '[']:
            if sep in var_info['long_name [units]']:
                long_name, units = [s.strip() for s in var_info['long_name [units]'].split(sep)]
                units = units.strip(']').strip(')').strip()
                var_info['long_name'] = long_name
                var_info['units'] = units
                break
        if 'long_name' not in var_info:
            var_info['long_name'] = var_info['long_name [units]']
        var_info.pop('long_name [units]')

# Ensure consistent info (i.e., columns) across all protocol tables (sheets)
columns = None
for key, var_info in protocol_vars.items():
    for s in ['out_name', 'units']:
        if s not in var_info:
            var_info[s] = None
    if columns is None:
        columns = set(var_info.keys())
    elif set(var_info.keys()) != columns:
        print(sorted(columns))
        print(sorted(var_info.keys()))
        raise Exception('inconsistent columns')

###############################################################################
# Load all CMOR tables to gather info about all available variables in already-defined projects.

# 'existing_cmorvars' dict will store full metadata for all variables from all projects,
# indexed by (project,table,variable) tuples.
existing_cmorvars = {}

for project, project_info in projects.items():
    path = project_info['path']
    ld = os.listdir(path)

    # Get table names
    # Assume filenames are '{project}_{table}.json'
    filenames = [s for s in ld if s.startswith(project) and s.endswith('.json') and s.count('_') == 1]
    tables = {}
    project_info.update({
        'tables' : tables,
        'no. of variables' : 0,
    })
    for filename in filenames:
        table = filename.split('_')[-1].replace('.json', '') # table name without the project name
        if table in project_info['exclude tables']:
            continue
        tables[table] = {
            'filename' : filename,
        }
    
    # Get info about tables
    for table, table_info in tables.items():
        filepath = os.path.join(path, table_info['filename'])
        with open(filepath, 'r') as f:
            dtab = json.load(f)
        if 'variable_entry' in dtab:
            table_info['variables'] = sorted(dtab['variable_entry'].keys())
            table_info['no. of variables'] = len(table_info['variables'])
            project_info['no. of variables'] += table_info['no. of variables']
            table_info['frequency'] = set()
            table_info['dimensions'] = {}
            for cmorvar, cmorvar_info in dtab['variable_entry'].items():
                variable_uid = (project, table, cmorvar) # tuple giving a unique id (uid) for the variable
                # should not be possible that this variable is already defined, but check just in case
                assert variable_uid not in existing_cmorvars, 'variable {} was already defined!'.format(variable_uid)
                existing_cmorvars[variable_uid] = cmorvar_info
                if 'frequency' in cmorvar_info:
                    freq = cmorvar_info['frequency']
                    table_info['frequency'].add(freq)
                    del freq
                if 'dimensions' in cmorvar_info:
                    dimensions = cmorvar_info['dimensions']
                    # Increment count of variables having these dimensions
                    if dimensions not in table_info['dimensions']:
                        table_info['dimensions'][dimensions] = 1
                    else:
                        table_info['dimensions'][dimensions] += 1
            table_info['frequency'] = sorted(table_info['frequency'])
        del dtab

print('Found total of {} variables in {} projects'.format(len(existing_cmorvars), len(projects)))
for project, project_info in projects.items():
    print('  {}: {}'.format(project, project_info['no. of variables']))

###############################################################################
# Go through QUOCA variables and attempt to find matches in CMOR tables of existing projects

for var_key, var_info in protocol_vars.items():  # loop over requested variables
    var_info.update({
        'matches' : [],
        'levels' : None,
    })
    for key, attrs in existing_cmorvars.items():  # loop over variables in already-existing project tables
        if 'frequency' not in attrs:
            continue

        conditions = []

        conditions.append(attrs['frequency'] == var_info['frequency'])

        if var_info['standard_name']:
            conditions.append(var_info['standard_name'] == attrs['standard_name'])
        elif var_info['out_name']:
            conditions.append(var_info['out_name'] == attrs['out_name'])
        else:
            conditions.append(key[-1] == var_info['variable_id'])

        spatial_shape = var_info['dimension']
        match spatial_shape:
            case '2-D zonal mean':
                conditions.append('longitude: mean' in attrs['cell_methods'])
            case '3-D':
                conditions.append('longitude' in attrs['dimensions'])
                conditions.append('latitude' in attrs['dimensions'])
                conditions.append('height' not in attrs['dimensions']) # indicates a surface field
            case '2-D (lat, lon)':
                conditions.append('longitude' in attrs['dimensions'])
                conditions.append('latitude' in attrs['dimensions'])
                conditions.append('plev' not in attrs['dimensions'])
            case _:
                raise Exception(f'Unknown spatial shape: {spatial_shape}')

        if all(conditions):
            var_info['matches'].append(key)

# print summary of how many matches for each variable
n = 0
for n in range(50):
    keys = [key for key, var_info in protocol_vars.items() if len(var_info['matches']) == n]
    nvar = len(keys)
    if nvar == 0:
        continue
    print(f'\nNo. of matches = {n}, {nvar} variables:')
    for key in sorted(keys):
        var_info = protocol_vars[key]
        w = f'  {key}'
        if n > 0:
            w += ' --> ' + ', '.join([str(t) for t in var_info['matches']])
        print(w)

###############################################################################
# Create CMOR tables and populate them with the best match for the variable
define_cmor_tables = {
    # define attributes matched in var_info
    'mon' : {
        'frequency' : ['mon'], 
        'dimension' : ['2-D (lat, lon)', '3-D'],
    },
    'monZ' : {
        'frequency' : ['mon'], 
        'dimension' : ['2-D zonal mean'],
    },
    'day' : {
        'frequency' : ['day'], 
        'dimension' : ['2-D (lat, lon)', '3-D'],
    },
    'dayZ' : {
        'frequency' : ['day'], 
        'dimension' : ['2-D zonal mean'],
    },
    '6hrPt' : {
        'frequency' : ['6hrPt'], 
        'dimension' : ['2-D (lat, lon)', '3-D'],
    },
    '6hrPtZ' : {
        'frequency' : ['6hrPt'], 
        'dimension' : ['2-D zonal mean'],
    },
}

project = 'QUOCA'
data_specs_version = "0.1.0"

filename_template = '{project}_{table_id}.json'
outpath = 'quoca-cmor-tables/Tables'

# order of preference to use when there are multiple matches:
use_project = ['QBOi', 'CCMI2022', 'SNAPSI', 'CMIP6']


cmor_table_header = {
    "data_specs_version": data_specs_version,
    "cmor_version": "3.5",
    # "table_id": "Table 6hrPtZ",
    "realm": "atmos",
    # "table_date": "September 19, 2022",
    "table_date" : datetime.datetime.now().strftime('%B %d, %Y'),
    "missing_value": "1e20",
    "int_missing_value": "-999",
    "product": "model-output",
    "approx_interval": "1.00000",
    "generic_levels": "",
    "mip_era": project, 
    "Conventions": f"CF-1.8 {project}"
}

protocol_table_levels = {
    'Table 2' : 'plev42',
    'Table 3' : 'plev42',
    'Table 4' : 'plevTEM',
    'Table 5' : 'plev42',
    'Table 6' : 'plevTEM',  # also: 15S to 15N, 150 to 0.4 hPa, 10 years of 1 ensemble member
}

print('\nCreating CMOR tables...')
assigned = set()
provenance = {}
for table_id, match_attrs in define_cmor_tables.items():
    keep = []
    for var_key, var_info in protocol_vars.items():
        conditions = []
        for attr in match_attrs:
            conditions.append(var_info[attr] in match_attrs[attr])
        if all(conditions):
            keep.append(var_key)

    if len(keep) == 0:
        print(f'no variables found for CMOR table {table_id}')
        continue

    cmor_table = {
        "Header" : dict(cmor_table_header),
        "variable_entry" : {},
    }
    cmor_table['Header']['table_id'] = table_id
    for var_key in keep:
        assert var_key not in assigned, f'{var_key} was already assigned to a CMOR table!'

        var_info = protocol_vars[var_key]
        variable_id = var_info['variable_id']
        assert variable_id not in cmor_table['variable_entry'] # variable_id must be a unique key within the CMOR table

        # uid = f'{variable_id}_{table_id}'
        uid = f'{project}_{table_id}_{variable_id}'
        assert uid not in provenance
        provenance[uid] = {
            'protocol table' : var_key[0],
            'spreadsheet variable name' : var_info['variable name in spreadsheet'],
            'spreadsheet long_name' : var_info['long_name'],
            'spreadsheet units' : var_info['units'],
            'spreadsheet dimension' : var_info['dimension'],
        }

        n = len(var_info['matches'])
        var_entry, key = None, None
        if n == 0:
            # no matches, so initialize a blank cmor variable
            var_entry = {
                "frequency": var_info['frequency'],
                "modeling_realm": "atmos", 
                "standard_name": "", 
                "units": "", 
                "cell_methods": "", 
                "cell_measures": "area: areacella",
                "long_name": var_info['long_name'], 
                "comment": "", 
                "dimensions": "", 
                "out_name": variable_id, # assume by default the out_name is the short variable name 
                "type": "real",  # seems reasonable
                "positive": "", 
                "valid_min": "", 
                "valid_max": "",
                "ok_min_mean_abs": "", 
                "ok_max_mean_abs": ""
            }
            for attr in ['standard_name', 'units']:
                if var_info[attr] is not None:
                    var_entry[attr] = var_info[attr]
        elif n == 1:
            # one match, no decision required!
            key = var_info['matches'][0]
        else:
            # multiple matches; pick one of them
            for proj in use_project:
                l = [key for key in var_info['matches'] if key[0] == proj]
                if len(l) > 0:
                    key = l[0]
                    break
        if key is None:
            provenance[uid]['source CMOR variable'] = 'none'
        else:
            var_entry = dict(existing_cmorvars[key])
            provenance[uid]['source CMOR variable'] = '_'.join(key)
        assert var_entry is not None
        del key

        # adjust levels if needed
        spatial_shape = var_info['dimension']
        levels = protocol_table_levels[var_key[0]]
        dimensions = None
        match spatial_shape:
            case '2-D zonal mean':
                dimensions = f'latitude {levels} time'
            case '3-D':
                dimensions = f'longitude latitude {levels} time'
            case '2-D (lat, lon)':
                dimensions = 'longitude latitude time'
            case _:
                raise Exception(f'Unknown spatial shape: {spatial_shape}')
        if 'cell_methods' in var_entry:
            if 'time: point' in var_entry['cell_methods']:
                dimensions = dimensions.replace('time', 'time1')
        assert dimensions is not None
        var_entry['dimensions'] = dimensions

        cmor_table['variable_entry'][variable_id] = var_entry

        assigned.add(var_key) # indicates that this variable has been added to a CMOR table

    if not os.path.exists(outpath):
        os.makedirs(outpath)
    filename = filename_template.format(table_id=table_id, project=project)
    filepath = os.path.join(outpath, filename)
    with open(filepath, 'w') as f:
        json.dump(cmor_table, f, indent=2)
        print('Wrote ' + filepath)

for var_key in protocol_vars:
    if var_key not in assigned:
        print(f'{var_key} was not assigned to any CMOR table!')

filename = f'{project}_cmorvars_provenance.json'
path = '.'
filepath = os.path.join(path, filename)
with open(filepath, 'w') as f:
    json.dump(provenance, f, indent=2)
    print('Wrote ' + filepath)

